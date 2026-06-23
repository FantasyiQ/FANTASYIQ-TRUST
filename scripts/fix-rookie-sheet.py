#!/usr/bin/env python3
"""
Clean a rookie spreadsheet's Player / School / Position cells using the seed
as the canonical source (correct proper-case, suffixes, acronyms) and trim
stray whitespace. Safer than heuristic title-casing — it copies the exact
strings already vetted in scripts/seed-rookie-rankings-<season>.ts.

Usage:
    python3 scripts/fix-rookie-sheet.py --season 2026

Matches rows to the seed by normalized name (case/punctuation-insensitive).
Rows with no seed match are left untouched and reported.
"""
import argparse, os, re, shutil, datetime
import openpyxl

def norm(s): return re.sub(r"[^a-z0-9]", "", str(s).lower())
def blank(v): return v in (None, "")

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--season", default="2026")
    ap.add_argument("--xlsx")
    ap.add_argument("--seed")
    args = ap.parse_args()
    xlsx = args.xlsx or os.path.expanduser(
        f"~/Library/CloudStorage/OneDrive-FantasyiQTrust/{args.season} NFL DRAFT.xlsx")
    seed = args.seed or os.path.join(os.path.dirname(__file__),
                                     f"seed-rookie-rankings-{args.season}.ts")

    # canonical name/school/position from the seed
    canon = {}
    for line in open(seed):
        if "playerName:" not in line:
            continue
        nm = re.search(r"playerName:\s*(?:'([^']*)'|\"((?:[^\"\\]|\\.)*)\")", line)
        sc = re.search(r"school:\s*'([^']*)'", line)
        po = re.search(r"position:\s*'([^']*)'", line)
        if not (nm and sc and po):
            continue
        name = nm.group(1) if nm.group(1) is not None else nm.group(2)
        canon[norm(name)] = (name, sc.group(1), po.group(1))
    print(f"Canonical entries from seed: {len(canon)}")

    bak = os.path.expanduser(
        f"~/Desktop/{args.season} NFL DRAFT.backup-{datetime.datetime.now():%Y%m%d-%H%M%S}.xlsx")
    shutil.copy(xlsx, bak)
    print(f"Backup: {bak}")

    wb = openpyxl.load_workbook(xlsx)
    ws = wb["Sheet1"]
    hdr = [c.value for c in ws[1]]
    ci = {h: i + 1 for i, h in enumerate(hdr)}

    changed = 0
    not_found = []
    for r in range(2, ws.max_row + 1):
        raw = ws.cell(row=r, column=ci["Player"]).value
        if blank(raw):
            continue
        key = norm(raw)
        if key not in canon:
            not_found.append(str(raw).strip())
            continue
        name, school, pos = canon[key]
        row_changed = False
        for col, val in (("Player", name), ("School", school), ("Position", pos)):
            if ws.cell(row=r, column=ci[col]).value != val:
                ws.cell(row=r, column=ci[col], value=val); row_changed = True
        # trim stray whitespace on Height
        h = ws.cell(row=r, column=ci["Height"]).value
        if isinstance(h, str) and h != h.strip():
            ws.cell(row=r, column=ci["Height"], value=h.strip()); row_changed = True
        if row_changed:
            changed += 1

    wb.save(xlsx)
    print(f"Cleaned {changed} rows.")
    if not_found:
        print(f"⚠️  {len(not_found)} sheet rows had no seed match (left untouched):")
        for n in not_found:
            print(f"     - {n}")

if __name__ == "__main__":
    main()
