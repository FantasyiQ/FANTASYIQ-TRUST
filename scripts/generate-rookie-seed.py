#!/usr/bin/env python3
"""
Generate a full rookie-rankings seed file from the master draft spreadsheet.

This turns "<SEASON> NFL DRAFT.xlsx" (Sheet1) into scripts/seed-rookie-rankings-<SEASON>.ts
— names, schools, positions, grades, scores, draft pick, measurements, and 40s —
so a new draft class is plug-and-play.

Usage:
    python3 scripts/generate-rookie-seed.py --season 2027
    npx tsx scripts/seed-rookie-rankings-2027.ts        # push to DB

Options:
    --season YYYY     Draft class year (default 2026)
    --xlsx PATH       Override spreadsheet path
    --out PATH        Override output .ts path
    --json PATH       Also write parsed player data as JSON (for validation)

Required Sheet1 columns: Player, School, Position, NFL Grade, FiQ Grade,
EliteScore, MarketScore, Pick (overall pick #), FiQScore, Height, Weight, 40.
draftCap is computed from Pick as round(100.2 - 0.2*Pick, 1); fiqTier is
computed in the seed's main() from the blended FiQ score.
"""
import argparse, json, math, os, re, subprocess, sys
import openpyxl

def norm_blank(v):
    return v in (None, "")

def ts_string(s: str) -> str:
    """Emit a TS string literal. Double-quote if it contains a single quote;
    escape any double quotes (e.g. inch marks in heights)."""
    s = str(s)
    if '"' in s:
        return '"' + s.replace('"', '\\"') + '"'
    if "'" in s:
        return '"' + s + '"'
    return "'" + s + "'"

def parse_forty(v):
    if norm_blank(v):
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None  # "DNP" / text => no verified 40

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--season", default="2026")
    ap.add_argument("--xlsx")
    ap.add_argument("--out")
    ap.add_argument("--json")
    ap.add_argument("--no-lint", action="store_true",
                    help="skip the sheet lint pre-check (not recommended)")
    args = ap.parse_args()

    season = args.season
    xlsx = args.xlsx or os.path.expanduser(
        f"~/Library/CloudStorage/OneDrive-FantasyiQTrust/{season} NFL DRAFT.xlsx")
    out = args.out or os.path.join(os.path.dirname(__file__), f"seed-rookie-rankings-{season}.ts")

    # Lint the sheet first — refuse to generate from a sheet with blocking errors.
    if not args.no_lint:
        lint = os.path.join(os.path.dirname(__file__), "lint-rookie-sheet.py")
        res = subprocess.run([sys.executable, lint, "--xlsx", xlsx],
                             capture_output=True, text=True)
        summary = next((l for l in res.stdout.splitlines() if l.startswith("Summary:")), "")
        if res.returncode > 0:
            print(res.stdout)
            print(f"\n⛔ Lint found {res.returncode} blocking error(s) — fix the sheet, "
                  f"or re-run with --no-lint to override.")
            sys.exit(1)
        print(f"Lint passed — {summary or 'clean'}")

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    idx = {h: i for i, h in enumerate(hdr)}
    need = ["Player", "School", "Position", "NFL Grade",
            "EliteScore", "MarketScore", "Pick", "Height", "Weight", 40]
    missing = [c for c in need if c not in idx]
    if missing:
        raise SystemExit(f"Spreadsheet missing required columns: {missing}")

    # ProspectGradeTable: NFL Grade (0.1 steps) -> FiQ Grade. fiqGrade is the
    # table value for the NFL grade floored to 0.1.
    grade_tbl = {}
    for gr in list(wb["ProspectGradeTable"].iter_rows(values_only=True))[1:]:
        if gr[0] is None or gr[1] is None:
            continue
        grade_tbl[round(float(gr[0]), 1)] = int(gr[1])

    def fiq_grade(nfl_grade: float):
        key = round(math.floor(nfl_grade * 10) / 10, 1)
        return grade_tbl.get(key)

    players = []
    skipped = []
    for r in rows[1:]:
        if not r or norm_blank(r[idx["Player"]]):
            continue
        name = str(r[idx["Player"]]).strip()
        pick = r[idx["Pick"]]
        if norm_blank(pick):
            skipped.append(f"{name} (no Pick)")
            continue
        # require the core input columns (FiQ grade/score are DERIVED, not read)
        req = ["NFL Grade", "EliteScore", "MarketScore"]
        if any(norm_blank(r[idx[c]]) for c in req):
            blanks = [c for c in req if norm_blank(r[idx[c]])]
            skipped.append(f"{name} (missing {', '.join(blanks)})")
            continue
        pick = int(pick)
        draft_cap = round(100.2 - 0.2 * pick, 1)
        nfl_grade = round(float(r[idx["NFL Grade"]]), 2)
        fiq_g = fiq_grade(nfl_grade)
        if fiq_g is None:
            skipped.append(f"{name} (NFL grade {nfl_grade} out of ProspectGradeTable range)")
            continue
        elite  = int(round(float(r[idx["EliteScore"]])))
        market = int(round(float(r[idx["MarketScore"]])))
        fiq_score = round(0.30 * fiq_g + 0.30 * elite + 0.10 * market, 1)
        p = {
            "playerName": name,
            "school":     str(r[idx["School"]]).strip(),
            "position":   str(r[idx["Position"]]).strip(),
            "nflGrade":   nfl_grade,
            "fiqGrade":   fiq_g,
            "eliteScore": elite,
            "marketScore":market,
            "overallPick":pick,
            "draftCap":   draft_cap,
            "fiqScore":   fiq_score,
            "height":     None if norm_blank(r[idx["Height"]]) else str(r[idx["Height"]]).strip(),
            "weight":     None if norm_blank(r[idx["Weight"]]) else int(r[idx["Weight"]]),
            "fortyTime":  parse_forty(r[idx[40]]),
        }
        players.append(p)

    # ── emit the .ts file ────────────────────────────────────────────────────
    lines = []
    for p in players:
        f = [
            f"playerName: {ts_string(p['playerName'])}",
            f"school: {ts_string(p['school'])}",
            f"position: {ts_string(p['position'])}",
            f"nflGrade: {p['nflGrade']:.2f}",
            f"fiqGrade: {p['fiqGrade']}",
            f"eliteScore: {p['eliteScore']}",
            f"marketScore: {p['marketScore']}",
            f"overallPick: {p['overallPick']}",
            f"draftCap: {p['draftCap']:.1f}",
            f"fiqScore: {p['fiqScore']:.1f}",
        ]
        if p["height"] is not None:
            f.append(f'height: {ts_string(p["height"])}')
        if p["weight"] is not None:
            f.append(f"weight: {p['weight']}")
        if p["fortyTime"] is not None:
            f.append(f"fortyTime: {p['fortyTime']:.2f}")
        lines.append("    { " + ", ".join(f) + " },")

    body = "\n".join(lines)
    content = f'''/**
 * Seed script — Dynasty Rookie Rankings {season}  (AUTO-GENERATED)
 *
 * Generated by scripts/generate-rookie-seed.py from "{season} NFL DRAFT.xlsx".
 * Do not hand-edit — re-run the generator after updating the spreadsheet.
 *
 * Run with:  npx tsx scripts/seed-rookie-rankings-{season}.ts
 * Idempotent: upserts on (season, playerName).
 */

import {{ PrismaClient }} from '@prisma/client';
import {{ PrismaPg }}    from '@prisma/adapter-pg';
import * as dotenv     from 'dotenv';
import * as path       from 'path';

dotenv.config({{ path: path.resolve(__dirname, '../.env.local') }});

const adapter = new PrismaPg({{ connectionString: process.env.DATABASE_URL! }});
const prisma  = new PrismaClient({{ adapter }});

const SEASON = '{season}';

const players = [
{body}
];

async function main() {{
    console.log(`\\nSeeding ${{players.length}} rookie rankings players for ${{SEASON}}...\\n`);

    let upserted = 0;
    for (const p of players) {{
        const baseFiQ  = parseFloat((p.fiqScore + p.draftCap * 0.25).toFixed(2));
        const fiqTier  = baseFiQ >= 85 ? 'Tier 1' : baseFiQ >= 78 ? 'Tier 2' : baseFiQ >= 70 ? 'Tier 3' : baseFiQ >= 62 ? 'Tier 4' : 'Tier 5';

        const extra = ('height' in p || 'weight' in p || 'fortyTime' in p) ? {{
            height:    (p as any).height    ?? null,
            weight:    (p as any).weight    ?? null,
            fortyTime: (p as any).fortyTime ?? null,
        }} : {{}};

        await prisma.rookieRankingsPlayer.upsert({{
            where:  {{ season_playerName: {{ season: SEASON, playerName: p.playerName }} }},
            update: {{
                school: p.school, position: p.position, nflGrade: p.nflGrade,
                fiqGrade: p.fiqGrade, eliteScore: p.eliteScore, marketScore: p.marketScore,
                overallPick: p.overallPick, draftCap: p.draftCap,
                baseFiQScore: baseFiQ, fiqScore: baseFiQ, fiqTier, ...extra,
            }},
            create: {{
                season: SEASON, playerName: p.playerName, school: p.school, position: p.position,
                nflGrade: p.nflGrade, fiqGrade: p.fiqGrade, eliteScore: p.eliteScore,
                marketScore: p.marketScore, overallPick: p.overallPick, draftCap: p.draftCap,
                baseFiQScore: baseFiQ, opportunityScore: 0, fiqScore: baseFiQ, fiqTier, ...extra,
            }},
        }});
        console.log(`  ✓  ${{String(++upserted).padStart(3)}}. ${{p.position.padEnd(4)}}  ${{p.playerName.padEnd(28)}}  FiQ ${{baseFiQ.toFixed(2).padStart(5)}}  ${{fiqTier}}`);
    }}
    console.log(`\\n✅  Done — ${{upserted}} players seeded for ${{SEASON}}.\\n`);
}}

main()
    .catch(e => {{ console.error(e); process.exit(1); }})
    .finally(() => prisma.$disconnect());
'''
    with open(out, "w") as fh:
        fh.write(content)
    if args.json:
        json.dump(players, open(args.json, "w"))

    print(f"Generated {out} with {len(players)} players (season {season}).")
    if skipped:
        print(f"⚠️  Skipped {len(skipped)} rows with no Pick:")
        for s in skipped:
            print(f"     - {s}")

if __name__ == "__main__":
    main()
