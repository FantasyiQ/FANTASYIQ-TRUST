#!/usr/bin/env python3
"""
Lint the rookie-rankings spreadsheet before generating the seed.

Flags problems that would corrupt or silently drop data during generation:
  ERRORS (block generation):
    - missing required columns / ProspectGradeTable tab
    - duplicate player names (would collide on upsert)
    - missing Pick, NFL Grade, EliteScore, or MarketScore (row gets skipped)
    - NFL Grade outside the ProspectGradeTable range
  WARNINGS (generate, but probably not what you want):
    - ALL-CAPS player name or school (would create uppercase/duplicate rows)
    - leading/trailing whitespace in Player/School/Position
    - 40 cell that isn't a number, blank, or "DNP" (the 40 would be dropped)
    - unknown position, missing/implausible weight, missing height

Usage:
    python3 scripts/lint-rookie-sheet.py --season 2027
    python3 scripts/lint-rookie-sheet.py --xlsx /path/to/sheet.xlsx

Exit code = number of ERRORS (0 = clean).
"""
import argparse, math, os, re, sys
import openpyxl

REQUIRED_COLS = ["Player", "School", "Position", "NFL Grade",
                 "EliteScore", "MarketScore", "Pick", "Height", "Weight", 40]
KNOWN_POS = {"QB","RB","WR","TE","FB","EDGE","DE","DT","NT","DL","LB","OLB","ILB","MLB",
             "CB","S","SAF","FS","SS","DB","K","P","LS"}
# legit all-caps school acronyms (won't be flagged)
ACRONYM_SCHOOLS = {"USC","TCU","LSU","UCF","SMU","BYU","UAB","UTEP","UTSA","UNLV",
                   "FIU","FAU","UCLA","UNC","UCONN","ULM","ECU","UNLV"}

def norm(s): return re.sub(r"[^a-z0-9]", "", str(s).lower())
def blank(v): return v in (None, "")

def is_all_caps(s):
    s = str(s)
    return bool(re.search(r"[A-Za-z]", s)) and s == s.upper() and s != s.lower()

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--season", default="2026")
    ap.add_argument("--xlsx")
    args = ap.parse_args()
    xlsx = args.xlsx or os.path.expanduser(
        f"~/Library/CloudStorage/OneDrive-FantasyiQTrust/{args.season} NFL DRAFT.xlsx")

    errors, warnings = [], []
    def err(msg): errors.append(msg)
    def warn(msg): warnings.append(msg)

    print(f"\nROOKIE SHEET LINT — {os.path.basename(xlsx)}\n" + "=" * 56)
    wb = openpyxl.load_workbook(xlsx, data_only=True)

    if "Sheet1" not in wb.sheetnames:
        print("ERROR: no 'Sheet1' tab."); sys.exit(1)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    idx = {h: i for i, h in enumerate(hdr)}

    # column / tab presence
    for c in REQUIRED_COLS:
        if c not in idx:
            err(f"missing required column: {c!r}")
    if "ProspectGradeTable" not in wb.sheetnames:
        err("missing 'ProspectGradeTable' tab (needed to derive FiQ Grade)")
    if errors:
        for e in errors: print(f"  ✗ ERROR  {e}")
        print("\nCannot lint rows until structure is fixed.")
        sys.exit(len(errors))

    grade_keys = set()
    for gr in list(wb["ProspectGradeTable"].iter_rows(values_only=True))[1:]:
        if gr[0] is not None:
            grade_keys.add(round(float(gr[0]), 1))

    seen = {}
    player_count = 0
    for ri, r in enumerate(rows[1:], start=2):
        if not r or blank(r[idx["Player"]]):
            continue
        player_count += 1
        raw_name = r[idx["Player"]]
        name = str(raw_name).strip()
        tag = f"row {ri:>3}  {name}"

        # duplicates
        k = norm(name)
        if k in seen:
            err(f"{tag} — duplicate of row {seen[k]} (upsert collision)")
        else:
            seen[k] = ri

        # whitespace
        for col in ("Player", "School", "Position"):
            v = r[idx[col]]
            if isinstance(v, str) and v != v.strip():
                warn(f"{tag} — leading/trailing space in {col}: {v!r}")

        # all-caps name
        if is_all_caps(raw_name):
            warn(f"{tag} — ALL-CAPS name (will generate uppercase)")
        sch = r[idx["School"]]
        if sch and is_all_caps(sch) and str(sch).strip().upper() not in ACRONYM_SCHOOLS \
           and (" " in str(sch).strip() or len(str(sch).strip()) >= 5):
            warn(f"{tag} — ALL-CAPS school: {str(sch).strip()!r}")

        # required numeric inputs
        if blank(r[idx["Pick"]]):
            err(f"{tag} — missing Pick (row would be skipped)")
        for col in ("NFL Grade", "EliteScore", "MarketScore"):
            if blank(r[idx[col]]):
                err(f"{tag} — missing {col} (row would be skipped)")

        # NFL grade in table range
        if not blank(r[idx["NFL Grade"]]):
            try:
                ng = float(r[idx["NFL Grade"]])
                key = round(math.floor(ng * 10) / 10, 1)
                if key not in grade_keys:
                    err(f"{tag} — NFL Grade {ng} (floor {key}) not in ProspectGradeTable")
            except (ValueError, TypeError):
                err(f"{tag} — NFL Grade not numeric: {r[idx['NFL Grade']]!r}")

        # position
        pos = str(r[idx["Position"]]).strip() if not blank(r[idx["Position"]]) else ""
        if pos and pos not in KNOWN_POS:
            warn(f"{tag} — unusual position: {pos!r}")

        # 40 sanity (non-number, non-DNP, non-blank => would be silently dropped)
        f = r[idx[40]]
        if not blank(f):
            try:
                float(f)
            except (ValueError, TypeError):
                if str(f).strip().upper() != "DNP":
                    warn(f"{tag} — 40 value {f!r} isn't a number or 'DNP' (will be dropped)")

        # weight / height
        w = r[idx["Weight"]]
        if blank(w):
            warn(f"{tag} — missing Weight")
        else:
            try:
                wv = int(w)
                if wv < 150 or wv > 400:
                    warn(f"{tag} — implausible Weight: {wv}")
            except (ValueError, TypeError):
                warn(f"{tag} — Weight not numeric: {w!r}")
        if blank(r[idx["Height"]]):
            warn(f"{tag} — missing Height")

    # ── report ────────────────────────────────────────────────────────────────
    if errors:
        print(f"\nERRORS ({len(errors)}) — block generation:")
        for e in errors: print(f"  ✗ {e}")
    if warnings:
        print(f"\nWARNINGS ({len(warnings)}):")
        for w in warnings: print(f"  ⚠ {w}")
    if not errors and not warnings:
        print("\n✅  Clean — no issues found.")
    print(f"\nSummary: {player_count} players, {len(errors)} errors, {len(warnings)} warnings.")
    sys.exit(len(errors))

if __name__ == "__main__":
    main()
